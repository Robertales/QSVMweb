import csv
import pandas as pd
from openpyxl import Workbook


def numberOfColumns(filename):
    """
    This function return the number of columns of a given dataset

    :param filename: name of the dataset
    :return: number of columns of the given dataset
    :rtype: int
    """

    f = open(filename, "r")
    reader = csv.reader(f, delimiter=",")
    numCols = len(next(reader))
    f.close()
    return numCols  # Read first line and count columns


def numberOfRows(filename):
    """
    This function return the number of raws of a given dataset

    :param filename: name of the dataset
    :return: number of raws of the given dataset
    :rtype: int
    """
    results = pd.read_csv(filename)

    return len(results)


def createFeatureList(numCols: int):
    """
    This function create a List of string such as ["feature1","feature2",...,"featureN"
    where N is given in input

    :param numCols: number of feature to insert in the list
    :return: List of feature
    :rtype: list
    """
    featureList = []
    for x in range(numCols):
        stringa = "feature{}".format(x + 1)
        featureList.append(stringa)
    return featureList


def writeTxt(fileName, list_values):
    """
    This function write in a file .txt the values stored in a list

    :param fileName: name of the output file
    :param list_values: list of values to write
    :return: name of the output file
    :rtype: str
    """
    f = open(fileName, "w+")
    for c in list_values:
        f.write(str(c) + "\n")
    f.close()
    return fileName


def writeXls(fileName, generations, evaluations, bestfits, times):
    """
    This function is used in Prototype Selection to write important info of the Genetic algorithm in a file .xls

    :param fileName: name of the output file
    :param generations:
    :param evaluations:
    :param bestfits:
    :param times:
    :return: name of the output file
    """
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Number of Generations")
    ws1.cell(column=2, row=1, value="Number of Evaluations")
    ws1.cell(column=3, row=1, value="Best Fitness Value")
    ws1.cell(column=4, row=1, value="Time in seconds")

    row = 2
    for g in generations:
        ws1.cell(column=1, row=row, value=g)
        row += 1

    row = 2
    for g in evaluations:
        ws1.cell(column=2, row=row, value=g)
        row += 1

    row = 2
    for g in bestfits:
        ws1.cell(column=3, row=row, value=g)
        row += 1

    row = 2
    for g in times:
        ws1.cell(column=4, row=row, value=g)
        row += 1

    wb.save(filename=fileName)
    wb.close()

    return fileName
